import openpyxl as xl
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow

# Authentication function
def get_authenticated_service():
    flow = InstalledAppFlow.from_client_secrets_file(
        'client_secrets.json',  # Replace with your JSON file
        scopes=['https://www.googleapis.com/auth/youtube.force-ssl']
    )
    credentials = flow.run_local_server(port=0)
    return build('youtube', 'v3', credentials=credentials)

# Function to fetch all stream keys and names
def get_stream_keys_and_names(youtube):
    streams = []
    request = youtube.liveStreams().list(
        part="id,cdn,snippet",
        mine=True
    )
    while request is not None:
        response = request.execute()
        for stream in response.get("items", []):
            streams.append({
                "Stream Name": stream["snippet"]["title"],
                "Stream Key": stream["cdn"]["ingestionInfo"]["streamName"],
                "Stream ID": stream["id"]
            })
        request = youtube.liveStreams().list_next(request, response)
    return streams

# Function to create a new live broadcast
def create_live_broadcast(youtube, title, scheduled_start_time, dvr, latency, autostart):
    broadcast_body = {
        "snippet": {
            "title": title,
            "scheduledStartTime": scheduled_start_time
        },
        "status": {
            "privacyStatus": "unlisted"  # Modify to "public" or "private" as needed
        },
        "contentDetails": {
            "enableDvr": dvr,
            "latencyPreference": latency,
            "enableAutoStart": autostart
        }
    }
    response = youtube.liveBroadcasts().insert(
        part="snippet,status,contentDetails",
        body=broadcast_body
    ).execute()
    return response["id"]

# Function to bind a stream to the broadcast
def bind_stream_to_broadcast(youtube, broadcast_id, stream_id):
    request = youtube.liveBroadcasts().bind(
        id=broadcast_id,
        part="id,contentDetails",
        streamId=stream_id
    )
    response = request.execute()
    return response

# Main function
def main():
    # Authenticate the YouTube API client
    youtube = get_authenticated_service()

    # Fetch all existing stream keys and names
    streams = get_stream_keys_and_names(youtube)

    # Load the Excel sheet
    workbook = xl.load_workbook("yt_details.xlsx")  # Replace with your file path
    sheet = workbook['7days']

    # Prepare output workbook
    output_workbook = xl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.append(["Broadcast Name", "Stream Key Name", "Broadcast ID", "YouTube Link", "Status"])

    row_index = 2
    while True:
        name = sheet.cell(row=row_index, column=2).value  # Skip the first column (language)
        scheduled_time = sheet.cell(row=row_index, column=3).value
        stream_key_name = sheet.cell(row=row_index, column=4).value
        dvr = sheet.cell(row=row_index, column=5).value  # "on" or "off" taken as is
        latency = sheet.cell(row=row_index, column=6).value  # Expected to be "normal", "low", "ultraLow", etc.
        autostart = sheet.cell(row=row_index, column=7).value  # "on" or "off" taken as is

        if not name or not scheduled_time or not stream_key_name:
            break

        # Find the corresponding Stream ID by stream key
        matching_stream = next((s for s in streams if s["Stream Key"] == stream_key_name), None)
        if not matching_stream:
            print(f"Error: No stream found with the key '{stream_key_name}'.")
            output_sheet.append([name, stream_key_name, None, None, "Error: Stream not found"])
            row_index += 1
            continue

        try:
            # Create the live broadcast
            broadcast_id = create_live_broadcast(
                youtube, name, scheduled_time, dvr == "on", latency, autostart == "on"
            )

            # Bind the broadcast to the stream
            bind_stream_to_broadcast(youtube, broadcast_id, matching_stream["Stream ID"])

            # Generate YouTube link
            youtube_link = f"https://www.youtube.com/watch?v={broadcast_id}"
            print(f"Successfully created and bound broadcast '{name}' to stream '{stream_key_name}'.")
            output_sheet.append([name, stream_key_name, broadcast_id, youtube_link, "Success"])

        except Exception as e:
            print(f"Error processing broadcast '{name}': {e}")
            output_sheet.append([name, stream_key_name, None, None, f"Error: {e}"])

        row_index += 1

    # Save the output workbook with the updated name
    output_workbook.save("output_links.xlsx")
    print("Broadcast details saved to 'output_links.xlsx'.")

if __name__ == "__main__":
    main()
